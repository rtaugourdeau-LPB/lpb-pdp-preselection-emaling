# -*- coding: utf-8 -*-
import io
import math
import re
import ssl
import smtplib
import unicodedata
import zipfile
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter

# ==============================================================================
# Configuration générale Streamlit (cloud-friendly)
# ==============================================================================
st.set_page_config(
    page_title="Présélection des projets",
    page_icon="📁",
    layout="centered",
)

# Dossier "virtuel" courant (juste pour info à l'écran)
try:
    BASE = Path(__file__).resolve().parent
except NameError:
    BASE = Path.cwd()

# ==============================================================================
# Constantes métier
# ==============================================================================
NOM_COL = "Nom"
VILLE_COL = "Ville"
PAYS_COL = "Pays"
DATE_COL_SUBMITTED = "Submitted At"
DATE_COL_FUNDS = "À quelle date souhaiteriez-vous obtenir ces fonds ?"

SUBFOLDERS_NA = ["Localisation non adaptée", "Non adapté"]

LABELS_RAW = (
    "Prénom\tNom\tNuméro de téléphone\tE-mail\tVous êtes ?\t"
    "S'agit-il de votre activité professionnelle principale ?  \t"
    "Avez-vous déjà réalisé des opérations à titre professionnel ?\t"
    "Parlons maintenant du projet qui vous conduit ici !\t"
    "Pouvez-vous nous décrire votre opération ?\tAdresse\tAdresse ligne 2\tVille\t"
    "État/Région/Province\tCode postal\tPays\t"
    "À date, sous quel état se trouve le foncier qui sert de base à votre opération ?\t"
    "Votre opération nécessite-t-elle une autorisation d'urbanisme (PA, DP, PC ...) ?\t"
    "De quel type(s) d'autorisation(s) avez-vous besoin pour réaliser votre opération ?\t"
    "Avez-vous déposé votre dossier en vue d'obtenir cette autorisation ?\t"
    "Bravo ! Et avez-vous obtenu cette autorisation ?\t"
    "Félicitations ! Et avez-vous déjà purgé cette autorisation ?\t"
    "Concernant la commercialisation de votre programme, l'avez-vous commencé ?\t"
    "À date, quel est votre pourcentage de commercialisation ?\t"
    "Quel est le coût de revient de cette opération ? \t"
    "Quel est le chiffre d'affaires que vous visez sur ce projet ?\t"
    "Une banque vous accompagne-t-elle sur ce projet ?\t"
    "Pouvez-vous nous préciser de quelle banque il s'agit et du montant de son financement ?\t"
    "Quel montant souhaitez-vous financer via La Première Brique ?\t"
    "À quelle date souhaiteriez-vous obtenir ces fonds ?\t"
    "Quel sera l'usage de ces fonds ?\t"
    "Quel est le montant des fonds propres que vous investissez dans ce projet ?\t"
    "Pour rentrer dans le détail, n'hésitez pas à charger votre bilan prévisionnel d'opération \t"
    "Un document de présentation si vous en avez un\t"
    "Et plus largement tout document que vous jugez utile !\t"
    "\"Une dernière question et c'est fini, promis !\nComment avez-vous connu La Première Brique ?\"\t"
    "J'autorise La Première Brique à conserver et utiliser les données transmises via ce formulaire\t"
    "\"Une dernière question et c'est fini, promis !\nComment avez-vous connu La Première Brique ?\"\t"
    "Nom de l'organisme de formation\tSubmitted At\tToken"
)
LABELS = LABELS_RAW.split("\t")

MONEY_LABELS = {
    "Quel est le coût de revient de cette opération ? ",
    "Quel est le chiffre d'affaires que vous visez sur ce projet ?",
    "Quel montant souhaitez-vous financer via La Première Brique ?",
    "Quel est le montant des fonds propres que vous investissez dans ce projet ?",
}

FR_MONTHS = [
    "Janvier",
    "Février",
    "Mars",
    "Avril",
    "Mai",
    "Juin",
    "Juillet",
    "Août",
    "Septembre",
    "Octobre",
    "Novembre",
    "Décembre",
]

# ==============================================================================
# Styles Excel
# ==============================================================================
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BOLD = Font(bold=True)
LEFT = Alignment(vertical="top")
WRAP = Alignment(wrap_text=True, vertical="top")
HEAD = PatternFill("solid", fgColor="E8F3FF")
ZEBRA = PatternFill("solid", fgColor="F7F7F7")

# ==============================================================================
# Fonctions utilitaires
# ==============================================================================
def slug(s: object) -> str:
    """Sanitise un texte pour l'utiliser dans un nom de fichier."""
    s = "" if pd.isna(s) else str(s)
    return re.sub(r'[\\/:\"*?<>|]+', " ", s).strip() or "INCONNU"


def robust_get(row: pd.Series, label: str):
    """Récupère une valeur dans la ligne en tolérant les variations d'espaces / casse."""
    if label in row.index:
        return row[label]
    norm = {re.sub(r"\s+", " ", str(c)).strip().casefold(): c for c in row.index}
    key = re.sub(r"\s+", " ", label).strip().casefold()
    return row.get(norm.get(key), None)


def format_money_text(x) -> str:
    """Formatage texte des montants pour l'Excel (espaces + €)."""
    if x is None:
        return ""
    s = str(x).strip().replace("\u202f", "")
    if "€" in s:
        return s
    s_no_sp = s.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"\d+(\.\d+)?", s_no_sp):
        if "." in s_no_sp:
            val = float(s_no_sp)
            entier = int(val)
            frac = int(round((val - entier) * 100))
            entier_f = f"{entier:,}".replace(",", " ")
            return f"{entier_f},{frac:02d} €"
        else:
            entier_f = f"{int(s_no_sp):,}".replace(",", " ")
            return f"{entier_f} €"
    return s


def month_label(dt: pd.Timestamp) -> str:
    """Libellé du mois type '10) Octobre - 2025'."""
    idx = int(dt.month) - 1
    prefix = f"{idx + 1}) {FR_MONTHS[idx]}"
    return f"{prefix} - {int(dt.year)}"


def width_from_texts(texts, extra=3) -> int:
    maxlen = 0
    for t in texts:
        if pd.isna(t):
            continue
        maxlen = max(maxlen, max(len(line) for line in str(t).split("\n")))
    return maxlen + extra


def title_fr(s: str) -> str:
    """Met une chaîne en 'Titre Français' (gestion apostrophes / tirets)."""
    if not s:
        return ""
    s = s.strip().lower().replace("’", "'")
    parts = re.split(r"([ \-'])", s)
    out = []
    for p in parts:
        if p in (" ", "-", "'") or p == "":
            out.append(p)
        else:
            out.append(p[:1].upper() + p[1:])
    return "".join(out)


def normalize_text(s: object) -> str:
    if not s:
        return ""
    s = "".join(
        c
        for c in unicodedata.normalize("NFD", str(s))
        if unicodedata.category(c) != "Mn"
    )
    s = s.strip().lower()
    s = re.sub(r"^[^a-z0-9]+", "", s)
    return s


def is_fr_country(pays, ville) -> bool:
    """Heuristique pour détecter si le projet est en France (métropole + DOM)."""
    s_pays = normalize_text(pays)
    s_ville = normalize_text(ville)
    domtom = {"re", "gp", "mq", "gf", "yt", "pm", "wf", "pf", "nc"}
    if (
        s_pays.startswith("fr")
        or s_pays.startswith("france")
        or s_pays in domtom
        or s_pays.startswith("republique fr")
    ):
        return True
    if not s_pays and s_ville.startswith("fr"):
        return True
    return False


def parse_date(val):
    """Parse robuste d'une date venant d'Excel ou de texte."""
    if isinstance(val, (pd.Timestamp,)):
        return pd.to_datetime(val)
    if isinstance(val, (int, float)) and not math.isnan(val):
        try:
            return pd.to_datetime(val, unit="d", origin="1899-12-30")
        except Exception:
            pass
    if val is None:
        return pd.NaT
    s = str(val).strip()
    if not s:
        return pd.NaT
    # 12/10-2025 → 12/10/2025
    s = re.sub(r"(\d{2}/\d{2})-(\d{4})$", r"\1/\2", s)
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return pd.to_datetime(s, errors="coerce")
    if re.match(r"^\d{2}/\d{2}/\d{4}", s):
        return pd.to_datetime(s, dayfirst=True, errors="coerce")
    return pd.NaT


def compute_month(row: pd.Series):
    """Retourne (clé_tri, libellé_mois) pour une ligne de formulaire."""
    d = parse_date(row.get(DATE_COL_SUBMITTED))
    if pd.isna(d):
        d = parse_date(row.get(DATE_COL_FUNDS))
    if pd.isna(d):
        return ((9999, 12, 31, row.name), "13) Sans date - À vérifier")
    return ((int(d.year), int(d.month), row.name), month_label(d))


def build_transposed_wb(row: pd.Series) -> Workbook:
    """Construit un Workbook Excel transposé (questions en colonne A, réponses en B)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transposition"
    a_texts, b_texts = [], []

    for i, label in enumerate(LABELS, start=1):
        raw = robust_get(row, label)
        text = "" if pd.isna(raw) else str(raw).strip()
        if isinstance(raw, bool):
            text = "Oui" if raw else "Non"
        if label in MONEY_LABELS:
            text = format_money_text(text)

        cA = ws.cell(i, 1, label)
        cA.font = BOLD
        cA.alignment = LEFT
        cA.border = BORDER
        cA.fill = HEAD

        cB = ws.cell(i, 2, text)
        cB.alignment = WRAP
        cB.border = BORDER

        if i % 2 == 0:
            cA.fill = ZEBRA
            cB.fill = ZEBRA

        a_texts.append(label)
        b_texts.append(text)

    ws.column_dimensions[get_column_letter(1)].width = width_from_texts(a_texts)
    ws.column_dimensions[get_column_letter(2)].width = width_from_texts(b_texts)
    return wb


def default_subject() -> str:
    return "Suivi de votre projet – localisation en dehors de notre zone d’intervention"


def default_body(prenom: str, nom: str) -> str:
    prenom = (prenom or "").strip().title()
    nom = (nom or "").strip().title()
    return f"""Bonjour {prenom} {nom},

Nous vous remercions pour l’intérêt porté à La Première Brique ainsi que pour la présentation de votre projet.

Après analyse attentive de votre dossier, nous sommes au regret de vous informer que nous ne serons pas en mesure d’y donner une suite favorable. En effet, notre plateforme intervient exclusivement sur des opérations situées en France métropolitaine. Or, votre projet est localisé à l’étranger, ce qui dépasse notre périmètre d’intervention.

Nous vous remercions néanmoins pour la confiance témoignée et vous souhaitons pleine réussite dans la réalisation de ce projet. 
Nous restons bien entendu à votre disposition pour étudier toute future opération conforme à notre zone géographique.

Bien cordialement,
L’équipe La Première Brique
"""


def send_email_smtp(
    user: str,
    password: str,
    to_addr: str,
    subject: str,
    body: str,
    host: str = "smtp.gmail.com",
    port: int = 465,
    use_ssl: bool = True,
):
    """Envoi SMTP basique (SSL ou STARTTLS)."""
    if not user or not password:
        raise RuntimeError("Identifiants email manquants.")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = to_addr
    msg.set_content(body)

    context = ssl.create_default_context()
    if use_ssl:
        with smtplib.SMTP_SSL(host, port, context=context) as server:
            server.login(user, password)
            server.send_message(msg)
    else:
        with smtplib.SMTP(host, port) as server:
            server.starttls(context=context)
            server.login(user, password)
            server.send_message(msg)


# ==============================================================================
# Initialisation de l'état de session
# ==============================================================================
if "contacts_non_adapte" not in st.session_state:
    st.session_state["contacts_non_adapte"] = []

if "excluded_contacts" not in st.session_state:
    st.session_state["excluded_contacts"] = set()

# ==============================================================================
# UI : navigation
# ==============================================================================
with st.sidebar:
    st.header("Menu")
    page = st.radio(
        "Aller à",
        ["Prérequis", "Transposition PDP", "Emailing"],
        index=1,
        label_visibility="collapsed",
    )

st.title("📁 Présélection des projets")
st.caption(f"Dossier courant (info) : {BASE}")

# ==============================================================================
# Page : Prérequis
# ==============================================================================
if page == "Prérequis":
    st.header("🧭 Prérequis")

    st.markdown(
        """
**Ce que fait l’outil :**

- 📄 **Transposition PDP** : transforme l’export Typeform en **1 Excel par porteur**, classé **par mois**, avec séparation France / hors France.
- ✉️ **Emailing** : liste automatiquement les projets **hors France** et permet d’envoyer les mails de refus (unitaire ou en masse).

⚙️ **Mode cloud :**

- Aucun fichier n’est lu/écrit sur le disque du serveur.
- Tout est géré **en mémoire** pendant votre session.
- Vous importez l’export PDP, l’app génère les Excel et un ZIP téléchargeable.
- Les projets *Localisation non adaptée* alimentent automatiquement l’onglet **Emailing**.
"""
    )

    st.divider()

    st.subheader("🔐 Pour l'envoi d’emails")
    st.markdown("""
    1) Crée un **mot de passe d’application** (type *Mail* / *Windows*).  
       👉 https://myaccount.google.com/apppasswords  
    2) Dans l'onglet **Emailing** une fois le fichier sélectionné entrez :
       - `email` = ton adresse Gmail  
       - `password` = **mot de passe d’application** (16 caractères, **sans espaces**)  
    """)

# ==============================================================================
# Page : Transposition PDP
# ==============================================================================
elif page == "Transposition PDP":
    st.header("📥 Transposition PDP (cloud)")

st.markdown(
    """
### 1) Importez l’export Typeform (Excel)
Données test : [Télécharger le fichier](https://lapremierebriquelpb-my.sharepoint.com/:x:/g/personal/r_taugourdeau_lapremierebrique_fr/IQDJJ1vtSZR5SJw8oTjp0LPVAVrFhT-pLWh9Zk1_Bk_Q69Y?e=L1aRZx)

### 2) Génération des fichiers
L’outil crée **un fichier Excel par porteur**, rangé par **mois**, et sépare :
- **France**
- **Localisation non adaptée**

### 3) Téléchargement
Vous pouvez télécharger un **ZIP**, et l’onglet *Emailing* est mis à jour.
"""
)


    uploaded = st.file_uploader(
        "Déposez l’Excel (ex: Réponse formulaire PDP.xlsx)",
        type=["xlsx"],
        help="Fichier d’export du formulaire PDP",
    )

    generate_zip = st.checkbox(
        "Générer un ZIP téléchargeable avec tous les fichiers", value=True
    )

    run = st.button(
        "Lancer la transposition",
        type="primary",
        disabled=(uploaded is None),
    )

    if run:
        if uploaded is None:
            st.error("Veuillez déposer un fichier Excel.")
            st.stop()

        try:
            df = pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"Erreur de lecture du fichier : {e}")
            st.stop()

        required_cols = {NOM_COL, "Prénom", "E-mail"}
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            st.error(
                "Colonnes obligatoires manquantes dans l’Excel : "
                + ", ".join(missing)
            )
            st.stop()

        pairs = df.apply(compute_month, axis=1, result_type="expand")
        df["_order"] = pairs[0]
        df["_month_label"] = pairs[1]
        df = df.sort_values("_order").copy()

        groups = list(df.groupby("_month_label", sort=False))
        nb_groups = len(groups)

        contacts_non_adapte = []
        files_for_zip = []
        total_rows = len(df)

        progress = st.progress(0)
        status = st.empty()

        for g_idx, (month_lbl, dfm) in enumerate(groups):
            month_dir_name = slug(month_lbl)

            rows_with_names = []
            for _, row in dfm.iterrows():
                nom = (str(row.get(NOM_COL, "")).strip().upper() or "INCONNU")
                ville_raw = row.get(VILLE_COL, "")
                ville = title_fr(
                    "" if pd.isna(ville_raw) else str(ville_raw).strip()
                )
                fname = f"{slug(nom)} - {slug(ville)}.xlsx"
                rows_with_names.append((fname, row))

            rows_with_names.sort(key=lambda t: t[0])

            for fname, row in rows_with_names:
                wb = build_transposed_wb(row)

                pays = row.get(PAYS_COL, "")
                ville_rule = (
                    ""
                    if pd.isna(row.get(VILLE_COL, ""))
                    else str(row.get(VILLE_COL, "")).strip()
                )
                is_fr = is_fr_country(pays, ville_rule)
                subdir = month_dir_name if is_fr else f"{month_dir_name}/Localisation non adaptée"

                bio = io.BytesIO()
                wb.save(bio)
                bio.seek(0)
                arcname = f"{subdir}/{slug(fname)}"

                if generate_zip:
                    files_for_zip.append((arcname, bio.getvalue()))

                if not is_fr:
                    prenom = str(row.get("Prénom", "")).strip().title()
                    nom_p = str(row.get("Nom", "")).strip().title()
                    email = str(row.get("E-mail", "")).strip()

                    contacts_non_adapte.append(
                        {
                            "id": f"{prenom}_{nom_p}_{email}_{month_lbl}_{fname}",
                            "prenom": prenom,
                            "nom": nom_p,
                            "email": email,
                            "month_label": month_lbl,
                            "file_name": fname,
                            "xlsx_bytes": bio.getvalue(),
                        }
                    )

            progress.progress(int((g_idx + 1) / max(1, nb_groups) * 100))
            status.info(f"Traitement du mois : {month_lbl}")

        st.session_state["contacts_non_adapte"] = contacts_non_adapte
        st.session_state["excluded_contacts"] = set()  # reset exclusions

        status.empty()
        st.success(
            f"✅ Transposition terminée : {total_rows} ligne(s) traitée(s), "
            f"{len(contacts_non_adapte)} projet(s) hors France détecté(s)."
        )

        if generate_zip and files_for_zip:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(
                zip_buffer, "w", compression=zipfile.ZIP_DEFLATED
            ) as z:
                seen = set()
                for arcname, content in files_for_zip:
                    if arcname in seen:
                        continue
                    seen.add(arcname)
                    z.writestr(arcname, content)
            zip_buffer.seek(0)

            st.download_button(
                "⬇️ Télécharger le ZIP des Excel générés",
                data=zip_buffer.getvalue(),
                file_name="Sorties_PDP_groupées_par_mois.zip",
                mime="application/zip",
            )

        if not contacts_non_adapte:
            st.info(
                "Aucun projet hors périmètre détecté. L’onglet Emailing sera vide pour cette importation."
            )
        else:
            st.info(
                "Les projets *Localisation non adaptée* sont maintenant disponibles dans l’onglet **Emailing**."
            )

# ==============================================================================
# Page : Emailing
# ==============================================================================
elif page == "Emailing":
    st.header("📧 Emailing — Projets hors périmètre")

    contacts = st.session_state.get("contacts_non_adapte", [])
    simulate = st.toggle(
        "Mode simulation (prévisualisation sans envoi réel)",
        value=True,
        help="En mode simulation, aucun email n'est envoyé au serveur SMTP.",
    )

    if not contacts:
        st.warning(
            "Aucun projet hors périmètre chargé. "
            "Allez d'abord dans l'onglet **Transposition PDP** et traitez un export."
        )
        st.stop()

    # ------------------------------
    # Paramètres SMTP saisis à la main
    # ------------------------------
    st.subheader("Paramètres SMTP")

    col_basic1, col_basic2 = st.columns(2)
    with col_basic1:
        smtp_user = st.text_input(
            "Adresse email expéditeur",
            value="",
            help="Adresse utilisée comme expéditeur (ex : no-reply@lapremierebrique.fr).",
        )
    with col_basic2:
        smtp_pass = st.text_input(
            "Mot de passe d'application / SMTP",
            type="password",
            help="Mot de passe d’application (Gmail) ou mot de passe SMTP fourni par votre IT.",
        )

    with st.expander("Paramètres SMTP avancés (à modifier seulement si besoin)"):
        col_adv1, col_adv2 = st.columns(2)
        with col_adv1:
            smtp_host = st.text_input(
                "Hôte SMTP",
                value="smtp.gmail.com",
                help="Ex : smtp.gmail.com, smtp.office365.com, smtp.votredomaine.fr",
            )
        with col_adv2:
            smtp_port = st.number_input(
                "Port",
                min_value=1,
                max_value=65535,
                value=465,
                step=1,
                help="465 pour SMTP SSL (Gmail), 587 pour STARTTLS.",
            )
        use_ssl = st.checkbox(
            "Utiliser SSL (recommandé pour le port 465)", value=True
        )

    EMAIL_CONFIGURED_UI = bool(smtp_user and smtp_pass and smtp_host and smtp_port)

    if not EMAIL_CONFIGURED_UI and not simulate:
        st.warning(
            "Renseignez au minimum l’adresse expéditeur, le mot de passe SMTP et l’hôte "
            "pour activer l’envoi réel. Sinon laissez le mode simulation activé."
        )

    # ------------------------------
    # Sélection des contacts / mois
    # ------------------------------
    all_months = sorted({c["month_label"] for c in contacts})
    now = datetime.now()

    def _default_month(months: list[str]) -> str | None:
        target = FR_MONTHS[now.month - 1]
        for m in months:
            if str(now.year) in m and target.lower() in m.lower():
                return m
        return months[-1] if months else None

    default_m = _default_month(all_months)

    if "months_multiselect" not in st.session_state:
        st.session_state["months_multiselect"] = (
            [default_m]
            if default_m in all_months
            else (all_months[-1:] if all_months else [])
        )

    def _set_months(value):
        st.session_state["months_multiselect"] = list(value)

    st.markdown("### 📆 Filtrer par mois")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.button("Tout sélectionner", on_click=_set_months, args=(all_months,))
    with c2:
        st.button("Tout désélectionner", on_click=_set_months, args=([],))
    with c3:
        st.button(
            "Mois courant",
            on_click=_set_months,
            args=([default_m] if default_m in all_months else [],),
        )

    selected_months = st.multiselect(
        "Choisis le ou les mois à afficher",
        options=all_months,
        default=st.session_state["months_multiselect"],
        key="months_multiselect",
        help="Les mois sont issus de la date du formulaire ou de la date souhaitée pour les fonds.",
    )

    filtered_contacts = [
        c for c in contacts if not selected_months or c["month_label"] in selected_months
    ]

    if not filtered_contacts:
        st.warning("Aucun contact à afficher avec les filtres actuels.")
        st.stop()

    st.info(
        f"📍 {len(filtered_contacts)} contact(s) hors périmètre dans les mois sélectionnés."
    )

    # ------------------------------
    # Liste des contacts avec édition unitaire
    # ------------------------------
    remaining_contacts = []

    for ctc in filtered_contacts:
        contact_id = ctc["id"]
        excluded = contact_id in st.session_state["excluded_contacts"]

        prenom = ctc["prenom"]
        nom = ctc["nom"]
        email = ctc["email"]
        fname = ctc["file_name"]
        month_label = ctc["month_label"]

        cols_header = st.columns([0.75, 0.25])
        with cols_header[0]:
            st.markdown(
                f"**📄 {fname}** — {prenom} {nom} &lt;{email or '—'}&gt;  "
                f"*(Mois : {month_label})*"
            )
        with cols_header[1]:
            if not excluded:
                if st.button(
                    "❌ Exclure de l'envoi",
                    key=f"exclude_{contact_id}",
                    help="Retirer ce contact de l'envoi global",
                ):
                    st.session_state["excluded_contacts"].add(contact_id)
                    st.rerun()
            else:
                if st.button(
                    "↩️ Réintégrer dans l'envoi",
                    key=f"include_{contact_id}",
                    help="Réintégrer ce contact dans l'envoi global",
                ):
                    st.session_state["excluded_contacts"].discard(contact_id)
                    st.rerun()

        if excluded:
            st.caption("🚫 Ce contact est actuellement exclu de l'envoi global.")
            st.divider()
            continue

        remaining_contacts.append(ctc)

        with st.expander(f"✉️ Détails du mail — {prenom} {nom}", expanded=False):
            to_key = f"to_{contact_id}"
            sub_key = f"sub_{contact_id}"
            body_key = f"body_{contact_id}"

            default_to = email
            default_sub = default_subject()
            default_msg = default_body(prenom, nom)

            to_edit = st.text_input("Destinataire", value=default_to, key=to_key)
            subject_edit = st.text_input("Objet", value=default_sub, key=sub_key)
            body_edit = st.text_area(
                "Message",
                value=default_msg,
                height=240,
                key=body_key,
            )

            st.markdown("---")
            center_col = st.columns([1, 2, 1])[1]
            with center_col:
                send_btn = st.button(
                    "🚨 Envoyer ce mail",
                    key=f"send_{contact_id}",
                    type="primary",
                    use_container_width=True,
                    disabled=simulate or not to_edit or not EMAIL_CONFIGURED_UI,
                )

            if simulate:
                st.info("Mode simulation activé : aucun mail ne sera envoyé.")
            elif not EMAIL_CONFIGURED_UI:
                st.error("Paramètres SMTP incomplets : envoi réel désactivé.")
            elif send_btn:
                try:
                    send_email_smtp(
                        smtp_user,
                        smtp_pass,
                        to_edit,
                        subject_edit,
                        body_edit,
                        host=smtp_host,
                        port=int(smtp_port),
                        use_ssl=use_ssl,
                    )
                    st.success(f"✅ Mail envoyé à {to_edit}")
                except Exception as e:
                    st.error(f"❌ Échec de l’envoi : {e}")

        st.divider()

    st.info(
        f"📩 {len(remaining_contacts)} contact(s) sélectionné(s) pour l'envoi global "
        f"(dont {len(st.session_state['excluded_contacts'])} exclu(s))."
    )

    # ------------------------------
    # Envoi global
    # ------------------------------
    st.markdown("---")
    center = st.columns([1, 2, 1])[1]
    with center:
        send_all = st.button(
            f"🚀 Tout envoyer ({len(remaining_contacts)})",
            type="primary",
            use_container_width=True,
            disabled=simulate or len(remaining_contacts) == 0 or not EMAIL_CONFIGURED_UI,
        )

    if send_all:
        st.session_state["show_confirm_all"] = True

    if st.session_state.get("show_confirm_all", False):
        st.warning(
            "⚠️ Êtes-vous sûr de **tout vouloir envoyer** ? "
            "Cette action enverra un email à chaque contact sélectionné (hors simulation)."
        )
        c_ok, c_cancel = st.columns(2)
        with c_ok:
            confirm_yes = st.button(
                "✅ Oui, envoyer maintenant", key="confirm_all_yes", type="primary"
            )
        with c_cancel:
            confirm_no = st.button("❌ Annuler", key="confirm_all_no")

        if confirm_no:
            st.session_state["show_confirm_all"] = False

        elif confirm_yes:
            st.session_state["show_confirm_all"] = False

            if simulate or not EMAIL_CONFIGURED_UI:
                st.info(
                    "Mode simulation ou paramètres SMTP incomplets : aucun mail envoyé."
                )
                st.stop()

            ok_count, err_count = 0, 0
            prog = st.progress(0)
            status = st.empty()

            total = max(1, len(remaining_contacts))
            for j, ctc in enumerate(remaining_contacts):
                cid = ctc["id"]
                prenom = ctc["prenom"]
                nom = ctc["nom"]

                to_addr = st.session_state.get(f"to_{cid}", ctc["email"])
                subject = st.session_state.get(f"sub_{cid}", default_subject())
                body = st.session_state.get(f"body_{cid}", default_body(prenom, nom))

                try:
                    if not to_addr:
                        err_count += 1
                        status.error(
                            f"⚠️ Pas d’email pour {prenom} {nom} — {ctc['file_name']}"
                        )
                    else:
                        send_email_smtp(
                            smtp_user,
                            smtp_pass,
                            to_addr,
                            subject,
                            body,
                            host=smtp_host,
                            port=int(smtp_port),
                            use_ssl=use_ssl,
                        )
                        ok_count += 1
                        status.success(f"✅ Envoyé à {prenom} {nom} ({to_addr})")
                except Exception as e:
                    err_count += 1
                    status.error(f"❌ Erreur sur {ctc['file_name']} : {e}")

                prog.progress(int((j + 1) / total * 100))

            st.success(
                f"🎯 Terminé : {ok_count} mail(s) envoyé(s), {err_count} erreur(s)."

            )






